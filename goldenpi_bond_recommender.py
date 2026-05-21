
"""
GoldenPi Bond Recommendation System
-----------------------------------
GoldenPi-only fixed-income recommender built around the user's exact policy:

1) Single platform availability
   - For GoldenPi-only mode this is inherently satisfied.

2) Time horizon matching
   - Prefer bonds closest to the client's target duration.

3) Listed securities preference
   - Prefer listed securities first.
   - GoldenPi sample file currently has no explicit listed/unlisted field, so the
     script keeps the rule structurally and treats all current rows as listed-unknown.

4) Yield optimization
   - After suitability filters, prioritize higher YTM.

5) Liquidity & issuer evaluation
   - Use rating / min lot / issue quality as tie-breakers.

6) Risk profile bucket logic
   - C = Conservative bucket
   - M = Moderate bucket
   - A = Aggressive bucket
   - Each client risk profile maps to a target allocation mix across C/M/A buckets.

The goal is portfolio construction, not just bond ranking.
"""

from __future__ import annotations

import argparse
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# User policy configuration
# ---------------------------------------------------------------------

PROFILE_BUCKET_WEIGHTS: Dict[str, Dict[str, float]] = {
    "Conservative": {"C": 0.60, "M": 0.20, "A": 0.20},
    "Moderate": {"C": 0.40, "M": 0.30, "A": 0.30},
    "Aggressive": {"C": 0.20, "M": 0.40, "A": 0.40},
}

# Rating score used for tie-breaks only.
RATING_SCORE = {
    "AAA": 100,
    "AA+": 92,
    "AA": 85,
    "AA-": 78,
    "A+": 70,
    "A": 62,
    "A-": 55,
    "BBB+": 45,
    "BBB": 35,
    "BBB-": 25,
    "BB+": 18,
    "BB": 12,
    "BB-": 8,
}

# Map ratings to the user's C/M/A bucket logic.
# You can adjust this mapping later if your internal policy changes.
RATING_TO_BUCKET = {
    "AAA": "C",
    "AA+": "C",
    "AA": "C",
    "AA-": "M",
    "A+": "M",
    "A": "M",
    "A-": "A",
    "BBB+": "A",
    "BBB": "A",
    "BBB-": "A",
    "BB+": "A",
    "BB": "A",
    "BB-": "A",
}

# Scores used for intra-bucket ranking.
BUCKET_SCORE = {"C": 100, "M": 70, "A": 40}

DEFAULT_PLANNING_DATE = pd.Timestamp.today().normalize()

MIN_LOT_RE = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*([LCc][rR]?|[Ll])\+?\s*$")
RATING_RE = re.compile(r"\b(AAA|AA\+|AA\-|AA|A\+|A\-|A|BBB\+|BBB\-|BBB|BB\+|BB\-|BB)\b", re.I)


# ---------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------

def parse_percent(value) -> Optional[float]:
    """
    Convert raw coupon / YTM values to percentage points.
    Examples:
        0.1125 -> 11.25
        9.65   -> 9.65
        "11.25%" -> 11.25
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None

    if isinstance(value, str):
        s = value.strip().replace("%", "")
        if not s:
            return None
        try:
            value = float(s)
        except ValueError:
            return None

    try:
        v = float(value)
    except Exception:
        return None

    if 0 < v < 1.5:
        return round(v * 100, 4)
    return round(v, 4)


def parse_date(value) -> Optional[pd.Timestamp]:
    """Parse dates and correct obvious 19xx maturity artefacts by adding 100 years."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None

    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return None

    if ts.year < 2000:
        try:
            ts = ts.replace(year=ts.year + 100)
        except ValueError:
            ts = ts + pd.DateOffset(years=100)

    return pd.Timestamp(ts).normalize()


def parse_amount_token(token: str) -> Optional[float]:
    """Parse amount strings like 5L, 10L+, 1Cr+, 20L+, 50L+ into INR."""
    if token is None:
        return None

    s = str(token).strip().replace(",", "")
    m = MIN_LOT_RE.match(s)
    if not m:
        return None

    num = float(m.group(1))
    unit = m.group(2).lower()

    if unit.startswith("cr"):
        return num * 10_000_000
    if unit.startswith("l"):
        return num * 100_000
    return None


def extract_rating(raw: str) -> Optional[str]:
    """Extract the worst (most conservative) rating from a combined rating string."""
    if raw is None or (isinstance(raw, float) and np.isnan(raw)):
        return None

    s = str(raw).upper()
    found = []
    for token in RATING_RE.findall(s):
        token = token.upper()
        if token in RATING_SCORE:
            found.append(token)

    if not found:
        return None

    # Choose the worst rating among those found.
    return max(found, key=lambda r: list(RATING_SCORE).index(r) if r in RATING_SCORE else 999)


def safe_float(value) -> Optional[float]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    try:
        return float(value)
    except Exception:
        return None


def bucket_from_rating(rating: Optional[str]) -> str:
    if not rating:
        return "A"
    return RATING_TO_BUCKET.get(rating, "A")


# ---------------------------------------------------------------------
# Load and normalize GoldenPi workbook
# ---------------------------------------------------------------------

def load_goldenpi_quotes(path: str) -> pd.DataFrame:
    """Read the GoldenPi workbook and return only data rows starting from the header row."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    for r in range(1, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if isinstance(first, str) and first.strip().upper() == "ISIN":
            header_row = r
            break

    if header_row is None:
        raise ValueError("Could not locate the GoldenPi header row containing 'ISIN'.")

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers, start=1)]

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if vals[0] is None:
            continue
        if not isinstance(vals[0], str):
            continue
        if not vals[0].startswith("INE"):
            continue
        rows.append(vals)

    if not rows:
        return pd.DataFrame(columns=headers[: ws.max_column])

    raw = pd.DataFrame(rows, columns=headers[: ws.max_column])
    return raw


def normalize_raw(raw: pd.DataFrame) -> pd.DataFrame:
    df = raw.copy()

    rename_map = {
        "ISIN": "isin",
        "Security": "security_name",
        "Face Value": "face_value",
        "Coupon": "coupon_raw",
        "Maturity Date": "maturity_date_raw",
        "Credit Rating": "credit_rating_raw",
        "Taxable": "taxable",
        "Last IP Date": "last_ip_date_raw",
        "Next IP Date": "next_ip_date_raw",
        "YTM/YTC": "ytm_raw",
        "Min lot": "min_lot_raw",
        "QTM": "qtm_raw",
    }
    df = df.rename(columns={c: rename_map.get(c, c) for c in df.columns})

    if "listed" not in df.columns:
        df["listed"] = None

    if "security_name" not in df.columns:
        raise ValueError("GoldenPi file is missing the 'Security' column.")

    df["coupon_pct"] = df.get("coupon_raw", pd.Series(index=df.index)).apply(parse_percent)
    df["ytm_pct"] = df.get("ytm_raw", pd.Series(index=df.index)).apply(parse_percent)
    df["maturity_date"] = df.get("maturity_date_raw", pd.Series(index=df.index)).apply(parse_date)
    df["last_ip_date"] = df.get("last_ip_date_raw", pd.Series(index=df.index)).apply(parse_date)
    df["next_ip_date"] = df.get("next_ip_date_raw", pd.Series(index=df.index)).apply(parse_date)
    df["credit_rating"] = df.get("credit_rating_raw", pd.Series(index=df.index)).apply(extract_rating)
    df["face_value"] = df.get("face_value", pd.Series(index=df.index)).apply(safe_float)
    df["min_lot_inr"] = df.get("min_lot_raw", pd.Series(index=df.index)).apply(parse_amount_token)
    df["qtm_inr"] = df.get("qtm_raw", pd.Series(index=df.index)).apply(parse_amount_token)

    taxable_source = df["taxable"] if "taxable" in df.columns else "Yes"
    df["taxable_flag"] = (
        taxable_source.astype(str)
        .str.strip()
        .str.lower()
        .map(lambda x: x in {"yes", "y", "true", "1"})
    )

    # GoldenPi sample does not include explicit listed/unlisted. Keep it as unknown/assumed listed.
    df["listed_flag"] = True

    df["bucket"] = df["credit_rating"].apply(bucket_from_rating)
    df["rating_score"] = df["credit_rating"].map(RATING_SCORE).fillna(0)

    df["security_name"] = df["security_name"].astype(str).str.strip()
    df["issuer"] = df["security_name"]  # no separate issuer field in current GoldenPi file

    today = DEFAULT_PLANNING_DATE
    df["days_to_maturity"] = (df["maturity_date"] - today).dt.days
    df["years_to_maturity"] = df["days_to_maturity"] / 365.25

    # Keep only usable instrument rows.
    df = df[df["isin"].astype(str).str.startswith("INE")].copy()

    return df.reset_index(drop=True)


# ---------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------

def duration_score(days_to_maturity: float, target_days: float) -> float:
    """Higher score when the bond maturity is close to the user's target horizon."""
    if days_to_maturity is None or pd.isna(days_to_maturity):
        return 0.0

    gap = abs(days_to_maturity - target_days)

    if gap <= 90:
        return 100.0
    if gap <= 180:
        return 85.0
    if gap <= 365:
        return 65.0
    if gap <= 540:
        return 45.0
    if gap <= 730:
        return 25.0
    return 10.0


def listed_score(listed_flag: Optional[bool]) -> float:
    """
    Listed securities should be prioritized first.
    In the current GoldenPi file this is unknown, so all rows are treated equally.
    """
    if listed_flag is True:
        return 100.0
    if listed_flag is False:
        return 60.0
    return 100.0


def liquidity_score(row: pd.Series) -> float:
    """
    Heuristic liquidity score using lot size and rating as proxies.
    Smaller lot sizes and better ratings improve liquidity.
    """
    score = 50.0

    lot = row.get("min_lot_inr")
    if lot is not None and not pd.isna(lot):
        if lot <= 500_000:
            score += 15
        elif lot <= 1_000_000:
            score += 8
        else:
            score -= 5

    rs = row.get("rating_score")
    if rs is not None and not pd.isna(rs):
        if rs >= 90:
            score += 20
        elif rs >= 80:
            score += 12
        elif rs >= 70:
            score += 5
        else:
            score -= 8

    return float(max(0, min(100, score)))


def issuer_score(row: pd.Series) -> float:
    rs = row.get("rating_score")
    if rs is None or pd.isna(rs):
        return 0.0
    return float(rs)


def yield_score(row: pd.Series, min_yield: float, max_yield: float) -> float:
    y = row.get("ytm_pct")
    if y is None or pd.isna(y):
        return 0.0
    if max_yield <= min_yield:
        return 50.0
    return float(((y - min_yield) / (max_yield - min_yield)) * 100.0)


def profile_weights(profile: str) -> Dict[str, float]:
    if profile not in PROFILE_BUCKET_WEIGHTS:
        raise ValueError(f"risk_profile must be one of {list(PROFILE_BUCKET_WEIGHTS)}")

    # Intra-bucket score weights. Yield comes after suitability, so it remains
    # a tie-breaker rather than the primary driver.
    if profile == "Conservative":
        return {"duration": 0.35, "listed": 0.15, "yield": 0.20, "liquidity": 0.15, "issuer": 0.15}
    if profile == "Moderate":
        return {"duration": 0.30, "listed": 0.15, "yield": 0.25, "liquidity": 0.15, "issuer": 0.15}
    return {"duration": 0.25, "listed": 0.10, "yield": 0.35, "liquidity": 0.15, "issuer": 0.15}


def bucket_preference_order(profile: str) -> list[str]:
    weights = PROFILE_BUCKET_WEIGHTS[profile]
    return sorted(weights.keys(), key=lambda b: (-weights[b], b))


# ---------------------------------------------------------------------
# Recommendation engine
# ---------------------------------------------------------------------

def recommend_bonds(
    df: pd.DataFrame,
    investment_amount: float,
    risk_profile: str,
    unlisted_allowed: bool,
    duration_months: int,
    top_n: int = 5,
) -> pd.DataFrame:
    """
    Rank bonds using the user's exact hierarchy and build a bucket-based recommendation set.
    """

    if risk_profile not in PROFILE_BUCKET_WEIGHTS:
        raise ValueError(f"risk_profile must be one of {list(PROFILE_BUCKET_WEIGHTS)}")

    work = df.copy()

    # Hard filters
    work = work[work["min_lot_inr"].notna()].copy()
    work = work[work["maturity_date"].notna()].copy()
    work = work[work["ytm_pct"].notna()].copy()
    work = work[work["min_lot_inr"] <= investment_amount].copy()

    if not unlisted_allowed and "listed_flag" in work.columns:
        work = work[work["listed_flag"] == True].copy()

    if work.empty:
        return work

    # Score all candidates; bucket assignment is based on rating.
    target_days = duration_months * 30.4375
    min_yield = work["ytm_pct"].min()
    max_yield = work["ytm_pct"].max()
    w = profile_weights(risk_profile)

    work["duration_score"] = work["days_to_maturity"].apply(lambda x: duration_score(x, target_days))
    work["listed_priority_score"] = work["listed_flag"].apply(listed_score)
    work["yield_score"] = work.apply(lambda r: yield_score(r, min_yield, max_yield), axis=1)
    work["liquidity_score"] = work.apply(liquidity_score, axis=1)
    work["issuer_score"] = work.apply(issuer_score, axis=1)

    # Bucket alignment score: the profile's target allocation mix decides the bucket importance.
    bucket_weight_map = PROFILE_BUCKET_WEIGHTS[risk_profile]
    work["bucket_weight"] = work["bucket"].map(bucket_weight_map).fillna(0.0)
    work["bucket_alignment_score"] = work["bucket"].map(BUCKET_SCORE).fillna(0.0)

    # Final score: suitability first, yield later.
    work["final_score"] = (
        w["duration"] * work["duration_score"]
        + w["listed"] * work["listed_priority_score"]
        + w["yield"] * work["yield_score"]
        + w["liquidity"] * work["liquidity_score"]
        + w["issuer"] * work["issuer_score"]
        + 0.05 * work["bucket_alignment_score"]
    )

    # Apply the user's stated preference order as tie-breakers.
    work = work.sort_values(
        by=[
            "bucket_weight",          # bucket allocation priority
            "final_score",            # composite suitability
            "days_to_maturity",       # closer horizon first
            "listed_priority_score",  # listed first
            "ytm_pct",                # yield optimization
            "liquidity_score",
            "issuer_score",
            "min_lot_inr",
        ],
        ascending=[False, False, True, False, False, False, False, True],
    ).reset_index(drop=True)

    cols = [
        "isin",
        "security_name",
        "bucket",
        "credit_rating",
        "coupon_pct",
        "ytm_pct",
        "maturity_date",
        "days_to_maturity",
        "min_lot_inr",
        "qtm_inr",
        "duration_score",
        "listed_priority_score",
        "yield_score",
        "liquidity_score",
        "issuer_score",
        "bucket_weight",
        "final_score",
    ]
    out = work[cols].copy()

    # Pretty formatting
    out["coupon_pct"] = out["coupon_pct"].map(lambda x: None if pd.isna(x) else round(x, 2))
    out["ytm_pct"] = out["ytm_pct"].map(lambda x: None if pd.isna(x) else round(x, 2))
    out["min_lot_inr"] = out["min_lot_inr"].map(lambda x: None if pd.isna(x) else int(x))
    out["qtm_inr"] = out["qtm_inr"].map(lambda x: None if pd.isna(x) else int(x))
    out["maturity_date"] = out["maturity_date"].dt.strftime("%Y-%m-%d")
    for c in ["duration_score", "listed_priority_score", "yield_score", "liquidity_score", "issuer_score", "bucket_weight", "final_score"]:
        out[c] = out[c].round(2)

    return out


def allocate_bucket_portfolio(
    ranked_df: pd.DataFrame,
    investment_amount: float,
    risk_profile: str,
    max_per_bond_pct: float = 0.40,
) -> pd.DataFrame:
    """
    Allocate capital according to the C/M/A mix for the chosen risk profile.
    Allocation is done bucket-by-bucket, then bond-by-bond inside each bucket.
    """
    if ranked_df.empty:
        return ranked_df

    target_weights = PROFILE_BUCKET_WEIGHTS[risk_profile]
    allocations = []
    remaining_total = investment_amount

    # Work per bucket in the exact portfolio mix order.
    for bucket in bucket_preference_order(risk_profile):
        bucket_df = ranked_df[ranked_df["bucket"] == bucket].copy()
        if bucket_df.empty:
            continue

        bucket_target = investment_amount * target_weights[bucket]
        remaining_bucket = bucket_target

        for _, row in bucket_df.iterrows():
            if remaining_bucket <= 0 or remaining_total <= 0:
                break

            lot = row["min_lot_inr"]
            if lot is None or pd.isna(lot) or lot <= 0:
                continue

            # Cap per bond, both relative to total investment and remaining bucket target.
            cap = min(investment_amount * max_per_bond_pct, remaining_bucket, remaining_total)

            # If the bucket target is smaller than the minimum lot, take one full lot
            # as long as the overall portfolio can support it.
            if cap < lot:
                if remaining_total >= lot:
                    alloc = lot
                else:
                    continue
            else:
                alloc = math.floor(cap / lot) * lot
                if alloc < lot and remaining_total >= lot:
                    alloc = lot

            allocations.append(
                {
                    **row.to_dict(),
                    "allocation_inr": int(alloc),
                    "lots": int(alloc // lot),
                }
            )
            remaining_bucket -= alloc
            remaining_total -= alloc

        # continue to next bucket even if current bucket still had unmet target,
        # because the user's policy allows using lower priority buckets as needed.

    # If cash remains, fill it using the highest-ranked remaining bonds regardless of bucket target.
    if remaining_total > 0 and ranked_df.shape[0] > 0:
        used_isins = {x["isin"] for x in allocations}
        leftovers = ranked_df[~ranked_df["isin"].isin(used_isins)].copy()
        for _, row in leftovers.iterrows():
            if remaining_total <= 0:
                break
            lot = row["min_lot_inr"]
            if lot is None or pd.isna(lot) or lot <= 0:
                continue
            cap = min(investment_amount * max_per_bond_pct, remaining_total)
            if cap < lot:
                if remaining_total >= lot:
                    alloc = lot
                else:
                    continue
            else:
                alloc = math.floor(cap / lot) * lot
                if alloc < lot and remaining_total >= lot:
                    alloc = lot

            allocations.append(
                {
                    **row.to_dict(),
                    "allocation_inr": int(alloc),
                    "lots": int(alloc // lot),
                }
            )
            remaining_total -= alloc

    alloc_df = pd.DataFrame(allocations)
    if alloc_df.empty:
        return alloc_df

    alloc_df = alloc_df.sort_values(
        by=["bucket_weight", "final_score", "ytm_pct"],
        ascending=[False, False, False],
    ).reset_index(drop=True)

    return alloc_df


# ---------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="GoldenPi bond recommender")
    p.add_argument("--file", required=True, help="Path to the GoldenPi Excel file")
    p.add_argument("--amount", type=float, required=True, help="Investment amount in INR")
    p.add_argument(
        "--risk",
        default="Moderate",
        choices=list(PROFILE_BUCKET_WEIGHTS.keys()),
        help="Risk profile",
    )
    p.add_argument(
        "--unlisted",
        default="yes",
        choices=["yes", "no"],
        help="Whether unlisted securities are allowed (GoldenPi file currently has no listed/unlisted field)",
    )
    p.add_argument("--duration", type=int, required=True, help="Target duration in months")
    p.add_argument("--top", type=int, default=10, help="Number of bonds to show in ranking")
    return p


def main():
    parser = build_arg_parser()
    args = parser.parse_args()

    raw = load_goldenpi_quotes(args.file)
    df = normalize_raw(raw)

    ranked = recommend_bonds(
        df=df,
        investment_amount=args.amount,
        risk_profile=args.risk,
        unlisted_allowed=(args.unlisted.lower() == "yes"),
        duration_months=args.duration,
        top_n=args.top,
    )

    print("\n=== GOLDENPI RECOMMENDATIONS ===")
    print(f"Investment Amount : ₹{args.amount:,.0f}")
    print(f"Risk Profile      : {args.risk}")
    print(f"Duration          : {args.duration} months")
    print(f"Raw Bonds Loaded  : {len(df)}")
    print(f"Eligible/Ranked   : {len(ranked)}")

    if ranked.empty:
        print("\nNo bonds matched your filters.")
        return

    print("\nTop Ranked Bonds:")
    print(ranked.head(args.top).to_string(index=False))

    alloc = allocate_bucket_portfolio(ranked, args.amount, args.risk)
    if alloc.empty:
        print("\nCould not construct an allocation with the available lots and constraints.")
        return

    print("\nSuggested Allocation by Bucket Logic:")
    cols = ["bucket", "security_name", "credit_rating", "allocation_inr", "lots", "min_lot_inr", "ytm_pct", "final_score"]
    print(alloc[cols].to_string(index=False))
    print(f"\nUnallocated Cash: ₹{args.amount - alloc['allocation_inr'].sum():,.0f}")

    # Show bucket totals
    bucket_summary = alloc.groupby("bucket", as_index=False)["allocation_inr"].sum()
    bucket_summary["pct"] = bucket_summary["allocation_inr"] / args.amount * 100
    print("\nBucket Allocation Summary:")
    print(bucket_summary.to_string(index=False))


if __name__ == "__main__":
    main()
